"""Bulk job import: parse TXT / CSV / XLSX uploads into job draft rows.

The importer only *parses* the upload into structured rows; validation (minimum
JD length) and persistence/analysis are handled by :class:`JobService`. Rows
that cannot be parsed are reported back to the caller instead of aborting the
whole import.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from app.shared.documents import normalize_text
from app.shared.errors import AppError, ErrorCode

SUPPORTED_IMPORT_EXTENSIONS = {".txt", ".md", ".csv", ".xlsx"}
MAX_IMPORT_BYTES = 5 * 1024 * 1024  # 5 MB
MAX_IMPORT_ROWS = 200

# Header aliases (lower-cased) for tabular imports.
_TITLE_KEYS = {"title", "岗位", "岗位名称", "职位", "职位名称"}
_COMPANY_KEYS = {"company", "公司", "公司名称"}
_CITY_KEYS = {"city", "城市", "工作地点", "地点"}
_JD_KEYS = {"jd", "jdtext", "jd_text", "jd text", "岗位描述", "职责", "描述", "jobdescription"}


@dataclass(slots=True)
class ParsedJob:
    title: str
    jd_text: str
    company: str | None = None
    city: str | None = None


@dataclass(slots=True)
class ImportParseResult:
    jobs: list[ParsedJob] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _extension(filename: str) -> str:
    name = filename.lower().strip()
    dot = name.rfind(".")
    return name[dot:] if dot != -1 else ""


def _pick(row: dict[str, str], keys: set[str]) -> str | None:
    for header, value in row.items():
        if header in keys and value:
            return value.strip() or None
    return None


def _rows_to_jobs(rows: list[dict[str, str]]) -> ImportParseResult:
    result = ImportParseResult()
    for index, raw in enumerate(rows, start=2):  # row 1 is the header
        row = {(k or "").strip().lower(): (v or "") for k, v in raw.items()}
        title = _pick(row, _TITLE_KEYS)
        jd_text = _pick(row, _JD_KEYS)
        if not title or not jd_text:
            result.errors.append(f"第 {index} 行缺少岗位名称或岗位描述，已跳过。")
            continue
        result.jobs.append(
            ParsedJob(
                title=title[:128],
                jd_text=normalize_text(jd_text),
                company=_pick(row, _COMPANY_KEYS),
                city=_pick(row, _CITY_KEYS),
            )
        )
    return result


def _parse_csv(data: bytes) -> ImportParseResult:
    text = data.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        return ImportParseResult(errors=["CSV 文件为空或缺少表头。"])
    return _rows_to_jobs(list(reader)[:MAX_IMPORT_ROWS])


def _parse_xlsx(data: bytes) -> ImportParseResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise AppError(
            ErrorCode.VALIDATION_ERROR, "服务器未启用 Excel 解析，请改用 CSV 导入。"
        ) from exc

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return ImportParseResult(errors=["Excel 文件没有可用的工作表。"])

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return ImportParseResult(errors=["Excel 文件为空或缺少表头。"])

    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
    dict_rows: list[dict[str, str]] = []
    for values in rows_iter:
        if values is None:
            continue
        record = {
            headers[i]: ("" if values[i] is None else str(values[i]))
            for i in range(min(len(headers), len(values)))
        }
        if any(v.strip() for v in record.values()):
            dict_rows.append(record)
        if len(dict_rows) >= MAX_IMPORT_ROWS:
            break
    workbook.close()
    return _rows_to_jobs(dict_rows)


def _parse_text(data: bytes) -> ImportParseResult:
    """A plain-text upload becomes a single job; title is the first line."""
    text = normalize_text(data.decode("utf-8", errors="ignore"))
    if not text:
        return ImportParseResult(errors=["文本文件为空。"])
    lines = text.split("\n")
    title = lines[0].strip()[:128] or "未命名岗位"
    return ImportParseResult(jobs=[ParsedJob(title=title, jd_text=text)])


def parse_jobs_from_file(filename: str, data: bytes) -> ImportParseResult:
    """Parse an uploaded file into job draft rows + per-row errors."""
    ext = _extension(filename)
    if ext not in SUPPORTED_IMPORT_EXTENSIONS:
        raise AppError(
            ErrorCode.VALIDATION_ERROR, "请上传 CSV、Excel(XLSX) 或 TXT 格式的岗位文件。"
        )
    if len(data) > MAX_IMPORT_BYTES:
        raise AppError(ErrorCode.VALIDATION_ERROR, "文件过大，请上传 5MB 以内的岗位文件。")

    if ext == ".csv":
        return _parse_csv(data)
    if ext == ".xlsx":
        return _parse_xlsx(data)
    return _parse_text(data)
